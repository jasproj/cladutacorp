#!/usr/bin/env python3
"""
claduta_drip.py — daily product-rotating cold-email drip for Claduta Corporation.

Phase-1 features (May 2026 rebuild):
  - Product rotation (coffee → sugar → pepper) + 15-day ramp [10..100]
  - Safety brake on 7-day failure-rate spike; mid-run check every 10 sends
  - Pre-flight forbidden-token guard (literal + regex)
  - Country filter (CASL + GDPR/EEA exclusions, hard block)
  - Role-address tiering (named > procurement > generic > other)
  - Domain throttling (1 send per recipient domain per day)
  - Idempotency keys (skip if email:product:date already sent today)
  - SendGrid suppression pre-fetch (bounces/blocks/spam/invalid), cached
  - Unsubscribe suppression list (populated by process_unsubscribes.py)
  - Single-instance lock so cron + manual runs cannot collide

Modes (mutually exclusive):
  --status   print queue + state, no SendGrid calls, no I/O
  --dry-run  print what WOULD send, uses cached SendGrid suppressions only
  --send     execute the send; fetches fresh SendGrid suppressions

All runtime artifacts (tracking CSV, summary CSV, state, logs, lock,
suppression caches) live in ~/cladutacorp/ regardless of where this
script is invoked from.
"""
from __future__ import annotations

import argparse
import csv
import fcntl
import json
import logging
import os
import re
import shutil
import sys
import time
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import requests
import warnings
warnings.filterwarnings("ignore", category=FutureWarning, module="google")
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────────────────
# Paths
# ─────────────────────────────────────────────────────────────────────────────

CLADUTA_HOME             = Path.home() / "cladutacorp"
TRACKING_CSV             = CLADUTA_HOME / "mega_blast_tracking.csv"
SUMMARY_CSV              = CLADUTA_HOME / "drip_summary.csv"
XLSX_PATH                = CLADUTA_HOME / "v11_updated.xlsx"
STATE_FILE               = CLADUTA_HOME / ".drip_state.json"
LOCK_FILE                = CLADUTA_HOME / ".drip.lock"
LOGS_DIR                 = CLADUTA_HOME / "logs"
UNSUBSCRIBE_CSV          = CLADUTA_HOME / "suppression_unsubscribes.csv"
SENDGRID_SUPPRESSION_CACHE = CLADUTA_HOME / ".sendgrid_suppressions.json"
TEMPLATES_DIR            = Path(__file__).resolve().parent / "drip_templates"
SHEETS_SA_CREDS          = Path.home() / ".secrets" / "claduta-sheets-service-account.json"

NAMED_CONTACTS_SHEET_ID = os.environ.get(
    "NAMED_CONTACTS_SHEET_ID",
    "1c4bFKmhSSbIA13ikFomQjMu4HZwiSxkrcJHFXSjmPdg",
)
NAMED_CONTACTS_TAB = os.environ.get("NAMED_CONTACTS_TAB", "Named Contacts")

PRODUCTS = ("coffee", "sugar", "pepper")

# ─────────────────────────────────────────────────────────────────────────────
# Sending policy
# ─────────────────────────────────────────────────────────────────────────────

COOLDOWN_DAYS             = 4
MIN_DAYS_BETWEEN_TOUCHES  = 30
MAX_TOUCHES_PER_EMAIL     = 3
BOUNCE_RATE_WINDOW_DAYS   = 7
BOUNCE_RATE_TRIP_PCT      = 5.0
SEND_PACING_SECONDS       = 10
HARD_DAILY_CAP            = 150
MID_RUN_BRAKE_INTERVAL    = 10
SENDGRID_CACHE_MAX_AGE_HR = 24.0
POOL_EXHAUSTION_ALERT_THRESHOLD = 10   # final queue below this ⇒ all pools drained

# 15-day ramp (replaces the 5-day ramp after external adversarial review).
# Indexed by state["ramp_day"]; clamps at the last value once we're warmed.
RAMP_SCHEDULE = [10, 10, 10, 15, 15, 15, 20, 20, 25, 25, 50, 50, 75, 75, 100, 125, 150]

XLSX_TABS_BY_PRODUCT: dict[str, list[str]] = {
    "coffee": ["Coffee", "Food Distributors"],
    "sugar":  ["Sugar — International", "Food Distributors"],
    "pepper": ["Black Pepper", "Food Distributors"],
}

NAMED_CONTACTS_PRODUCT_MAP: dict[str, str] = {
    "Coffee": "coffee",
    "Coffee — International": "coffee",
    "Sugar — International": "sugar",
    "Sugar": "sugar",
    "Black Pepper": "pepper",
    "Pepper": "pepper",
    "Pepper — International": "pepper",
    # "Food Distributors" handled specially: aliased to today's rotation product
}

# ─────────────────────────────────────────────────────────────────────────────
# Forbidden-token guard
# ─────────────────────────────────────────────────────────────────────────────

FORBIDDEN_GLOBAL = [
    "Sacconi", "Ibanez", "Alexandre", "AA7", "COOABRIEL",
    "Lucas Silva", "Matheus Silva", "Sacconi Pimentas",
]
FORBIDDEN_BY_PRODUCT: dict[str, list[str]] = {
    "coffee": ["Santos"],   # coffee NEVER ships from Santos
    "pepper": ["Santos"],   # pepper FOB Vitória only
    "sugar":  [],           # sugar legitimately uses Santos
}
FORBIDDEN_REGEX_BY_PRODUCT: dict[str, list[str]] = {
    "pepper": [r"\$\s*\d"], # NO pepper pricing in cold outbound
}

# ─────────────────────────────────────────────────────────────────────────────
# Country filter (ITEM B). Hard-block Canada (CASL) + EU/UK/EEA (GDPR — no
# documented LIA). Match on full name OR ISO2 code OR email TLD fallback.
# ─────────────────────────────────────────────────────────────────────────────

EXCLUDED_COUNTRY_NAMES = {
    "canada",
    "united kingdom", "uk", "great britain", "england", "scotland", "wales",
    "ireland", "france", "germany", "italy", "spain", "netherlands", "holland",
    "belgium", "luxembourg", "austria", "denmark", "sweden", "finland",
    "portugal", "greece", "poland", "czech republic", "czechia", "slovakia",
    "slovenia", "hungary", "romania", "bulgaria", "croatia", "estonia",
    "latvia", "lithuania", "malta", "cyprus", "switzerland", "norway",
    "iceland", "liechtenstein",
}
EXCLUDED_COUNTRY_ISO2 = {
    "CA", "UK", "GB", "IE", "FR", "DE", "IT", "ES", "NL", "BE", "LU",
    "AT", "DK", "SE", "FI", "PT", "GR", "PL", "CZ", "SK", "SI", "HU",
    "RO", "BG", "HR", "EE", "LV", "LT", "MT", "CY", "CH", "NO", "IS", "LI",
}
EXCLUDED_TLDS = {
    ".ca", ".uk", ".gb", ".ie", ".fr", ".de", ".it", ".es", ".nl", ".be",
    ".lu", ".at", ".dk", ".se", ".fi", ".pt", ".gr", ".pl", ".cz", ".sk",
    ".si", ".hu", ".ro", ".bg", ".hr", ".ee", ".lv", ".lt", ".mt", ".cy",
    ".ch", ".no", ".is", ".li", ".eu",
}

# ─────────────────────────────────────────────────────────────────────────────
# Role-address tiering (ITEM C)
# ─────────────────────────────────────────────────────────────────────────────

TIER1_LOCALPARTS = {
    "procurement", "buying", "sourcing", "purchasing", "imports", "trading",
}
TIER2_LOCALPARTS = {
    "info", "sales", "contact", "admin", "hello", "office", "support",
    "enquiries", "inquiries",
}

# ─────────────────────────────────────────────────────────────────────────────
# SendGrid suppression sync (ITEM D)
# ─────────────────────────────────────────────────────────────────────────────

SENDGRID_SUPPRESSION_ENDPOINTS = [
    ("bounces",        "https://api.sendgrid.com/v3/suppression/bounces"),
    ("blocks",         "https://api.sendgrid.com/v3/suppression/blocks"),
    ("spam_reports",   "https://api.sendgrid.com/v3/suppression/spam_reports"),
    ("invalid_emails", "https://api.sendgrid.com/v3/suppression/invalid_emails"),
]
SENDGRID_API_URL = "https://api.sendgrid.com/v3/mail/send"

# ─────────────────────────────────────────────────────────────────────────────
# Scanner health (BLOCKER 2). Drip refuses --send when the unsubscribe
# scanner is flagged unhealthy or its last_ok marker is stale (> 36h),
# so we never blast through a missed-opt-out window.
# ─────────────────────────────────────────────────────────────────────────────

SCANNER_UNHEALTHY_FILE        = CLADUTA_HOME / ".unsubscribe_scanner_unhealthy"
SCANNER_LAST_OK_FILE          = CLADUTA_HOME / ".unsubscribe_scanner_last_ok"
SCANNER_LAST_OK_MAX_AGE_HOURS = 36.0

# ─────────────────────────────────────────────────────────────────────────────
# CSV schemas
# ─────────────────────────────────────────────────────────────────────────────

TRACKING_COLS = [
    "email", "company", "product", "city",
    "sent_at",
    "followup1_sent_at", "followup2_sent_at",
    "status",
    "followup3_sent_at",
]
# "failed" preserved as backward-compat aggregate (rejected + api_errors).
SUMMARY_COLS = ["date", "product", "attempted", "succeeded", "rejected",
                "api_errors", "failed", "bounce_rate_7d"]

# ─────────────────────────────────────────────────────────────────────────────
# Custom exceptions
# ─────────────────────────────────────────────────────────────────────────────

class SendGridSuppressionAbort(RuntimeError):
    """All SendGrid suppression endpoints failed and the cache is stale/absent."""

class TemplateGuardFailed(RuntimeError):
    """Pre-flight forbidden-token check found violations."""

# ─────────────────────────────────────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────────────────────────────────────

def setup_logger(run_id: str) -> logging.Logger:
    LOGS_DIR.mkdir(exist_ok=True)
    log_path = LOGS_DIR / f"drip_{run_id}.log"
    fmt = logging.Formatter("[%(asctime)s] %(message)s", datefmt="%Y-%m-%dT%H:%M:%S")
    logger = logging.getLogger("drip")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fh = logging.FileHandler(log_path); fh.setFormatter(fmt)
    sh = logging.StreamHandler(sys.stdout); sh.setFormatter(fmt)
    logger.addHandler(fh); logger.addHandler(sh)
    return logger

# ─────────────────────────────────────────────────────────────────────────────
# Single-instance lock
# ─────────────────────────────────────────────────────────────────────────────

_lock_fp = None

def acquire_send_lock(logger: logging.Logger) -> bool:
    global _lock_fp
    CLADUTA_HOME.mkdir(exist_ok=True)
    _lock_fp = open(LOCK_FILE, "w")
    try:
        fcntl.flock(_lock_fp, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except BlockingIOError:
        logger.error(f"Another drip instance holds the lock at {LOCK_FILE}. Aborting.")
        return False
    _lock_fp.write(str(os.getpid()))
    _lock_fp.flush()
    return True

# ─────────────────────────────────────────────────────────────────────────────
# State
# ─────────────────────────────────────────────────────────────────────────────

def load_state() -> dict:
    if not STATE_FILE.exists():
        return {
            "last_product_sent": None,
            "ramp_day": 0,
            "last_send_date": None,
            "safety_brake": {"active": False, "reason": None, "triggered_at": None},
        }
    return json.loads(STATE_FILE.read_text())

def save_state(s: dict) -> None:
    CLADUTA_HOME.mkdir(exist_ok=True)
    STATE_FILE.write_text(json.dumps(s, indent=2))

def next_product(last_product: str | None) -> str:
    if last_product not in PRODUCTS:
        return PRODUCTS[0]
    i = PRODUCTS.index(last_product)
    return PRODUCTS[(i + 1) % len(PRODUCTS)]

def ramp_batch_size(ramp_day: int) -> int:
    if ramp_day < 0:
        return RAMP_SCHEDULE[0]
    if ramp_day >= len(RAMP_SCHEDULE):
        return RAMP_SCHEDULE[-1]
    return RAMP_SCHEDULE[ramp_day]

# ─────────────────────────────────────────────────────────────────────────────
# Tracking CSV I/O
# ─────────────────────────────────────────────────────────────────────────────

def _parse_dt(s: str | None) -> datetime | None:
    if not s:
        return None
    s = s.strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None

# Status classification (BLOCKER 4). The send loop writes one of:
#   SENT          — HTTP 2xx
#   REJECTED      — HTTP 4xx (bad address, blocked) — counts toward bounce rate
#   API_THROTTLED — HTTP 429                       — does NOT count
#   API_ERROR     — HTTP 5xx, timeout, network err — does NOT count
# Legacy mega_blast_tracking.csv rows used status="FAILED"; we treat those
# as REJECTED for the bounce math (best guess at legacy semantics).
def _status_is_sent(status: str) -> bool:
    return (status or "").strip().upper() == "SENT"

def _status_is_bounce(status: str) -> bool:
    """Recipient-side rejection. Excludes transient API failures."""
    s = (status or "").strip().upper()
    if s in ("API_ERROR", "API_THROTTLED"):
        return False
    if s == "FAILED":                       # legacy mega_blast rows
        return True
    return any(tok in s for tok in ("REJECTED", "BOUNCED", "DEAD", "COMPLAINED"))

def _status_counts_for_reputation(status: str) -> bool:
    return _status_is_sent(status) or _status_is_bounce(status)


def load_tracking() -> tuple[dict[str, list[datetime]], set[str], dict[str, int]]:
    touches: dict[str, list[datetime]] = defaultdict(list)
    bounced: set[str] = set()
    touch_count: dict[str, int] = defaultdict(int)
    if not TRACKING_CSV.exists():
        return touches, bounced, touch_count
    with open(TRACKING_CSV, newline="") as f:
        for row in csv.DictReader(f):
            email = (row.get("email") or "").strip().lower()
            if not email:
                continue
            for col in ("sent_at", "followup1_sent_at", "followup2_sent_at", "followup3_sent_at"):
                ts = _parse_dt(row.get(col, ""))
                if ts:
                    touches[email].append(ts)
                    touch_count[email] += 1
            if _status_is_bounce(row.get("status", "")):
                bounced.add(email)
    return touches, bounced, touch_count

def _atomic_csv_append(path: Path, fieldnames: list[str], row: dict) -> None:
    """Atomic append (BLOCKER 3). Copy existing → .tmp, append + fsync,
    os.replace. Doubles I/O per write; acceptable at 10–100/day. Caller
    guarantees single-instance access (drip lock; summary is post-loop)."""
    CLADUTA_HOME.mkdir(exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    if path.exists():
        shutil.copy2(path, tmp)
        mode = "a"
        need_header = False
    else:
        if tmp.exists():
            tmp.unlink()
        mode = "w"
        need_header = True
    with open(tmp, mode, newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if need_header:
            writer.writeheader()
        writer.writerow({c: row.get(c, "") for c in fieldnames})
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, path)

def append_tracking_row(row: dict) -> None:
    _atomic_csv_append(TRACKING_CSV, TRACKING_COLS, row)

def append_summary_row(row: dict) -> None:
    _atomic_csv_append(SUMMARY_CSV, SUMMARY_COLS, row)

def compute_7d_failure_rate() -> float:
    """Send-time failure rate = REJECTED / (SENT + REJECTED) over the 7-day window.

    Counts ONLY the drip's own send-time outcomes:
      - SENT     (HTTP 2xx)            -> denominator
      - REJECTED (HTTP 4xx bad addr)   -> numerator + denominator
    EXCLUDES from the rate entirely:
      - API_ERROR / API_THROTTLED      (transient — SendGrid health, not the recipient)
      - BOUNCED / DEAD / COMPLAINED / legacy FAILED (retroactive labels applied by
        SendGrid reconciliation, NOT a send attempt — so relabeling old bounces must
        not trip the brake).
    The brake therefore fires only when the drip is actively getting send-time
    rejections (i.e. sending to bad addresses right now). Already-bounced addresses
    are still suppressed from future sends via load_tracking()'s bounced set —
    that path is unchanged."""
    if not TRACKING_CSV.exists():
        return 0.0
    cutoff = datetime.now() - timedelta(days=BOUNCE_RATE_WINDOW_DAYS)
    counted = rejected = 0
    with open(TRACKING_CSV, newline="") as f:
        for row in csv.DictReader(f):
            ts = _parse_dt(row.get("sent_at", ""))
            if not ts or ts < cutoff:
                continue
            status = (row.get("status", "") or "").strip().upper()
            if status == "SENT":
                counted += 1
            elif status == "REJECTED":
                counted += 1
                rejected += 1
            # all other statuses excluded from the rate (see docstring)
    if counted == 0:
        return 0.0
    return (rejected / counted) * 100.0

def _recent_send_emails(within_days: int = 7) -> set[str]:
    """Emails (lowercased) with a sent_at row in the window."""
    if not TRACKING_CSV.exists():
        return set()
    cutoff = datetime.now() - timedelta(days=within_days)
    out: set[str] = set()
    with open(TRACKING_CSV, newline="") as f:
        for row in csv.DictReader(f):
            ts = _parse_dt(row.get("sent_at", ""))
            if not ts or ts < cutoff:
                continue
            em = (row.get("email") or "").strip().lower()
            if em:
                out.add(em)
    return out

def load_recent_send_keys(within_hours: int = 24) -> set[tuple[str, str, str]]:
    """ITEM F idempotency: {(email_lower, product_lower, YYYY-MM-DD)} for sends
    in the rolling window. Compared against today's sends to skip dupes."""
    if not TRACKING_CSV.exists():
        return set()
    cutoff = datetime.now() - timedelta(hours=within_hours)
    keys: set[tuple[str, str, str]] = set()
    with open(TRACKING_CSV, newline="") as f:
        for row in csv.DictReader(f):
            ts = _parse_dt(row.get("sent_at", ""))
            if not ts or ts < cutoff:
                continue
            em = (row.get("email") or "").strip().lower()
            prod = (row.get("product") or "").strip().lower()
            if em:
                keys.add((em, prod, ts.strftime("%Y-%m-%d")))
    return keys

# ─────────────────────────────────────────────────────────────────────────────
# Country filter (ITEM B)
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_country(s: str) -> str:
    s = s.lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def country_excluded(country_str: str | None) -> bool:
    if not country_str:
        return False
    norm = _normalize_country(country_str)
    if not norm:
        return False
    for name in EXCLUDED_COUNTRY_NAMES:
        if name in norm:
            return True
    for tok in norm.split():
        if tok.upper() in EXCLUDED_COUNTRY_ISO2:
            return True
    return False

def tld_excluded(email: str) -> bool:
    em = email.lower().rstrip()
    for tld in EXCLUDED_TLDS:
        if em.endswith(tld):
            return True
    return False

def apply_country_filter(candidates: list[dict], logger: logging.Logger) -> tuple[list[dict], int]:
    """
    Drop contacts in excluded countries. Priority: contact['country'] →
    contact['city'] (xlsx Region/Country lives here for schema-compat) →
    email-TLD fallback (backstop only).
    """
    kept: list[dict] = []
    excluded = 0
    samples: list[str] = []
    for c in candidates:
        country = (c.get("country") or c.get("city") or "").strip()
        hit = country_excluded(country)
        if not hit and not country:
            hit = tld_excluded(c.get("email", ""))
        if hit:
            excluded += 1
            if len(samples) < 5:
                samples.append(f"{c.get('email')} ({country or 'tld'})")
            continue
        kept.append(c)
    if excluded:
        logger.info(f"country_filter: excluded {excluded} contact(s); sample={samples}")
    return kept, excluded

# ─────────────────────────────────────────────────────────────────────────────
# Role-address tier (ITEM C)
# ─────────────────────────────────────────────────────────────────────────────

def email_tier(email: str, first_name: str | None) -> int:
    """0 = named human, 1 = procurement role, 2 = generic inbox, 3 = other."""
    if first_name and first_name.strip():
        return 0
    local = email.split("@", 1)[0].lower() if "@" in email else email.lower()
    local = local.split("+", 1)[0]  # strip "+suffix"
    if local in TIER1_LOCALPARTS:
        return 1
    if local in TIER2_LOCALPARTS:
        return 2
    return 3

# ─────────────────────────────────────────────────────────────────────────────
# Unsubscribe suppression list (ITEM A consumer)
# Populated by process_unsubscribes.py — script reads CSV, merges into bounced.
# ─────────────────────────────────────────────────────────────────────────────

def load_unsubscribe_suppressions() -> set[str]:
    if not UNSUBSCRIBE_CSV.exists():
        return set()
    out: set[str] = set()
    try:
        with open(UNSUBSCRIBE_CSV, newline="") as f:
            for row in csv.DictReader(f):
                e = (row.get("email") or "").strip().lower()
                if e:
                    out.add(e)
    except Exception:
        pass
    return out

# ─────────────────────────────────────────────────────────────────────────────
# SendGrid suppression sync (ITEM D)
# ─────────────────────────────────────────────────────────────────────────────

def fetch_sendgrid_suppressions(api_key: str, logger: logging.Logger
                                ) -> tuple[set[str], dict[str, int], int]:
    """Returns (all_emails, per_endpoint_counts, num_failed_endpoints).
    Per-endpoint failures are logged and counted; caller decides what to do
    if all four failed."""
    headers = {"Authorization": f"Bearer {api_key}"}
    all_emails: set[str] = set()
    counts: dict[str, int] = {}
    failed = 0
    PAGE = 500          # SendGrid v3 suppression list max page size
    MAX_PAGES = 400     # runaway guard (200k records) — far beyond any real list
    for label, url in SENDGRID_SUPPRESSION_ENDPOINTS:
        n = 0
        offset = 0
        endpoint_failed = False
        for _page in range(MAX_PAGES):
            try:
                r = requests.get(
                    url,
                    headers=headers,
                    params={"limit": PAGE, "offset": offset},
                    timeout=20,
                )
                r.raise_for_status()
                data = r.json()
            except Exception as e:
                logger.warning(
                    f"sendgrid_suppressions[{label}] fetch failed at offset={offset}: {e}"
                )
                endpoint_failed = True
                break
            page_records = len(data) if isinstance(data, list) else 0
            if isinstance(data, list):
                for item in data:
                    em = (item.get("email") or "").strip().lower()
                    if em:
                        all_emails.add(em)
            n += page_records
            # Last page reached when the API returns fewer than a full page.
            if page_records < PAGE:
                break
            offset += PAGE
        # Only treat as a failed endpoint if we got nothing at all; a mid-run
        # page error keeps the records already accumulated.
        if endpoint_failed and n == 0:
            counts[label] = -1
            failed += 1
        else:
            counts[label] = n
    return all_emails, counts, failed

def save_sendgrid_cache(emails: set[str], counts: dict[str, int]) -> None:
    CLADUTA_HOME.mkdir(exist_ok=True)
    SENDGRID_SUPPRESSION_CACHE.write_text(json.dumps({
        "ts": datetime.now().isoformat(),
        "emails": sorted(emails),
        "counts": counts,
    }, indent=2))

def load_sendgrid_cache() -> tuple[set[str], float | None, dict[str, int]]:
    """Returns (emails, age_hours_or_None, counts)."""
    if not SENDGRID_SUPPRESSION_CACHE.exists():
        return set(), None, {}
    try:
        d = json.loads(SENDGRID_SUPPRESSION_CACHE.read_text())
        ts = datetime.fromisoformat(d["ts"])
        age_h = (datetime.now() - ts).total_seconds() / 3600.0
        return set(d.get("emails", [])), age_h, d.get("counts", {})
    except Exception:
        return set(), None, {}

# ─────────────────────────────────────────────────────────────────────────────
# Domain throttling (ITEM E)
# ─────────────────────────────────────────────────────────────────────────────

def domain_throttle(eligible: list[dict]) -> tuple[list[dict], int]:
    """Keep first occurrence per recipient domain. Relies on `eligible` being
    pre-sorted by priority (tier asc, then touch_count asc)."""
    seen: set[str] = set()
    out: list[dict] = []
    for c in eligible:
        email = c["email"].strip().lower()
        domain = email.split("@", 1)[1] if "@" in email else email
        if domain in seen:
            continue
        seen.add(domain)
        out.append(c)
    return out, len(eligible) - len(out)

# ─────────────────────────────────────────────────────────────────────────────
# Scanner health gate (BLOCKER 2)
# ─────────────────────────────────────────────────────────────────────────────

def get_scanner_health() -> dict:
    """Inspect the unsubscribe-scanner marker files. Returns:
        state:   HEALTHY | UNHEALTHY | STALE | NEVER_RAN
        healthy: bool (if False, --send must refuse)
        reason:  str | None
        last_ok_age_hours: float | None
        warnings: list[str]   (printable lines for status/dry-run output)
    """
    out: dict = {
        "state": "HEALTHY",
        "healthy": True,
        "reason": None,
        "last_ok_age_hours": None,
        "warnings": [],
    }
    if SCANNER_UNHEALTHY_FILE.exists():
        try:
            reason = SCANNER_UNHEALTHY_FILE.read_text().strip()[:300]
        except Exception:
            reason = "(unreadable)"
        out["state"]   = "UNHEALTHY"
        out["healthy"] = False
        out["reason"]  = reason
        out["warnings"].append(f"unsubscribe scanner UNHEALTHY: {reason}")
        return out
    if not SCANNER_LAST_OK_FILE.exists():
        out["state"]   = "NEVER_RAN"
        out["healthy"] = False
        out["warnings"].append("unsubscribe scanner has never recorded a successful run")
        return out
    try:
        mtime = datetime.fromtimestamp(SCANNER_LAST_OK_FILE.stat().st_mtime)
        age_h = (datetime.now() - mtime).total_seconds() / 3600.0
        out["last_ok_age_hours"] = age_h
        if age_h > SCANNER_LAST_OK_MAX_AGE_HOURS:
            out["state"]   = "STALE"
            out["healthy"] = False
            out["warnings"].append(
                f"unsubscribe scanner last OK {age_h:.1f}h ago "
                f"(>{SCANNER_LAST_OK_MAX_AGE_HOURS}h threshold)"
            )
    except Exception as e:
        out["state"]   = "UNHEALTHY"
        out["healthy"] = False
        out["reason"]  = f"could not read last_ok marker: {e}"
        out["warnings"].append(out["reason"])
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Contact sources
# ─────────────────────────────────────────────────────────────────────────────

def load_xlsx_contacts(product: str, logger: logging.Logger) -> list[dict]:
    if not XLSX_PATH.exists():
        logger.warning(f"xlsx not found at {XLSX_PATH} — skipping xlsx source")
        return []
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    out: list[dict] = []
    for tab in XLSX_TABS_BY_PRODUCT[product]:
        if tab not in wb.sheetnames:
            logger.warning(f"xlsx tab not found: {tab!r} (skipping)")
            continue
        ws = wb[tab]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            continue
        idx = {h: i for i, h in enumerate(header)}
        i_email   = idx.get("Email")
        i_company = idx.get("Company")
        i_region  = idx.get("Region/Country") or idx.get("State")
        i_status  = idx.get("Status")
        if i_email is None:
            logger.warning(f"xlsx tab {tab!r} has no Email column — skipping")
            continue
        for row in rows:
            email_raw = row[i_email]
            if not email_raw:
                continue
            email = str(email_raw).strip()
            if "@" not in email:
                continue
            status = ""
            if i_status is not None and row[i_status]:
                status = str(row[i_status]).upper()
            if "DEAD" in status or "BOUNC" in status:
                continue
            company = str(row[i_company]).strip() if i_company is not None and row[i_company] else ""
            region  = str(row[i_region]).strip()  if i_region  is not None and row[i_region]  else ""
            out.append({
                "email":      email,
                "company":    company,
                "city":       region,   # preserves legacy tracking-CSV column
                "country":    region,   # used by ITEM B country filter
                "product":    product,
                "first_name": None,
                "source":     f"xlsx:{tab}",
            })
    logger.info(f"xlsx: {len(out)} rows for product={product}")
    return out

def load_named_contacts(product: str, logger: logging.Logger) -> list[dict]:
    if not SHEETS_SA_CREDS.exists():
        logger.warning(f"Sheets SA creds missing at {SHEETS_SA_CREDS} — skipping Named Contacts")
        return []
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError as e:
        logger.warning(f"gspread / google-auth not installed ({e}) — skipping Named Contacts")
        return []
    try:
        creds = Credentials.from_service_account_file(
            str(SHEETS_SA_CREDS),
            scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
        )
        client = gspread.authorize(creds)
        sh = client.open_by_key(NAMED_CONTACTS_SHEET_ID)
        ws = sh.worksheet(NAMED_CONTACTS_TAB)
        records = ws.get_all_records()
    except Exception as e:
        logger.warning(f"Named Contacts read failed: {e}")
        return []
    out: list[dict] = []
    for r in records:
        email = (r.get("Email") or "").strip()
        if not email or "@" not in email:
            continue
        sheet_label = (r.get("Sheet") or "").strip()
        if sheet_label == "Food Distributors":
            mapped = product
        else:
            mapped = NAMED_CONTACTS_PRODUCT_MAP.get(sheet_label)
        if mapped != product:
            continue
        first = (r.get("First Name") or "").strip() or None
        # Defensive: Named Contacts tab does not currently carry a Country
        # column, but if added later the country filter will pick it up.
        country = (r.get("Country") or "").strip()
        out.append({
            "email":      email,
            "company":    (r.get("Company") or "").strip(),
            "city":       country,
            "country":    country,
            "product":    product,
            "first_name": first,
            "source":     "sheet:Named Contacts",
        })
    logger.info(f"named-contacts: {len(out)} rows for product={product}")
    return out

# Legacy mega-blast recipients live in TRACKING_CSV, NOT in the XLSX / Named
# Contacts pool, so the drip never followed them up (3,476 of ~3,705 were
# invisible). Re-surface them as candidates here. The standard pipeline still
# gates every send: apply_country_filter, SendGrid + unsubscribe suppression,
# touch_count cap, and the 30-day cooldown in filter_eligible. Product mapping:
# real commodity tags pass through; directory-scraped tags (foodex / europages
# / none / blank) → sugar (broadest offering, per operator decision 2026-06-13).
TRACKING_FOLLOWUP_PRODUCT_MAP = {
    "coffee":       "coffee",
    "sugar":        "sugar",
    "pepper":       "pepper",
    "black pepper": "pepper",   # explicit pepper buyers — keep on pepper, not the sugar default
    "spice":        "pepper",   # spice-directory traders → pepper (operator decision 2026-06-13)
}
TRACKING_FOLLOWUP_DEFAULT_PRODUCT = "sugar"

def load_tracking_followup_contacts(product: str, logger: logging.Logger) -> list[dict]:
    """Candidates drawn from prior-blast recipients in TRACKING_CSV that the
    XLSX / Named-Contacts pool does not contain. Each row's legacy product tag
    is mapped to a drip rotation product; only rows mapping to `product` are
    returned. Deduped by email. Eligibility (touch cap, cooldown), suppression,
    and country exclusion are enforced downstream — this only surfaces them."""
    if not TRACKING_CSV.exists():
        return []
    out: list[dict] = []
    seen: set[str] = set()
    with open(TRACKING_CSV, newline="") as f:
        for row in csv.DictReader(f):
            email = (row.get("email") or "").strip()
            if not email or "@" not in email:
                continue
            # Only re-surface addresses that actually received an initial send.
            if not (row.get("sent_at") or "").strip():
                continue
            tag = (row.get("product") or "").strip().lower()
            mapped = TRACKING_FOLLOWUP_PRODUCT_MAP.get(tag, TRACKING_FOLLOWUP_DEFAULT_PRODUCT)
            if mapped != product:
                continue
            key = email.lower()
            if key in seen:
                continue
            seen.add(key)
            region = (row.get("city") or "").strip()   # legacy col; read by country filter
            out.append({
                "email":      email,
                "company":    (row.get("company") or "").strip(),
                "city":       region,
                "country":    region,
                "product":    product,
                "first_name": None,
                "source":     "tracking:followup",
            })
    logger.info(f"tracking-followup: {len(out)} row(s) for product={product}")
    return out

# ─────────────────────────────────────────────────────────────────────────────
# Eligibility filter (with tier annotation, sorted by tier then touch_count)
# ─────────────────────────────────────────────────────────────────────────────

def filter_eligible(
    candidates: list[dict],
    touches: dict[str, list[datetime]],
    bounced: set[str],
    touch_count: dict[str, int],
    now: datetime,
) -> list[dict]:
    eligible: list[dict] = []
    seen: set[str] = set()
    for c in candidates:
        email = c["email"].strip().lower()
        if not email or email in seen:
            continue
        seen.add(email)
        if email in bounced:
            continue
        if touch_count.get(email, 0) >= MAX_TOUCHES_PER_EMAIL:
            continue
        if email in touches and touches[email]:
            last = max(touches[email])
            age = now - last
            if age < timedelta(days=COOLDOWN_DAYS):
                continue
            if age < timedelta(days=MIN_DAYS_BETWEEN_TOUCHES):
                continue
        c["tier"] = email_tier(email, c.get("first_name"))
        eligible.append(c)
    # ITEM C sort: tier asc, then touch_count asc (0-touch precedes 1, 2…)
    eligible.sort(key=lambda c: (c["tier"],
                                 touch_count.get(c["email"].strip().lower(), 0)))
    return eligible

# ─────────────────────────────────────────────────────────────────────────────
# Templates
# ─────────────────────────────────────────────────────────────────────────────

SUBJECT_LINE_RE = re.compile(r"^SUBJECT:\s*(.+?)\s*\n", re.IGNORECASE)

def load_templates(product: str) -> list[tuple[str, str, str]]:
    tdir = TEMPLATES_DIR / product
    if not tdir.exists():
        raise FileNotFoundError(f"Template directory not found: {tdir}")
    out: list[tuple[str, str, str]] = []
    for f in sorted(tdir.glob("variant_*.txt")):
        text = f.read_text()
        m = SUBJECT_LINE_RE.match(text)
        if not m:
            raise ValueError(f"Template {f.name} missing 'SUBJECT: ...' on first line")
        subject = m.group(1).strip()
        body = text[m.end():].lstrip("\n")
        out.append((f.stem, subject, body))
    if not out:
        raise FileNotFoundError(f"No variant_*.txt files found in {tdir}")
    return out

def forbidden_token_check(product: str, templates: list[tuple[str, str, str]]) -> list[str]:
    violations: list[str] = []
    literal_forbid = FORBIDDEN_GLOBAL + FORBIDDEN_BY_PRODUCT.get(product, [])
    regex_forbid = [re.compile(p) for p in FORBIDDEN_REGEX_BY_PRODUCT.get(product, [])]
    for vid, subj, body in templates:
        for tok in literal_forbid:
            if tok in subj or tok in body:
                violations.append(f"{product}/{vid}: contains forbidden token {tok!r}")
        for pat in regex_forbid:
            for fname, field in (("subject", subj), ("body", body)):
                m = pat.search(field)
                if m:
                    violations.append(
                        f"{product}/{vid}: {fname} matches forbidden regex "
                        f"{pat.pattern!r} (matched {m.group()!r})"
                    )
    return violations

def render(subject: str, body: str, contact: dict) -> tuple[str, str]:
    fn = (contact.get("first_name") or "").strip()
    rendered_body = body.replace("{{first_name}}", fn if fn else "there")
    return subject, rendered_body

# ─────────────────────────────────────────────────────────────────────────────
# SendGrid send
# ─────────────────────────────────────────────────────────────────────────────

def send_via_sendgrid(api_key: str, from_email: str, from_name: str,
                      to_email: str, subject: str, body: str,
                      reply_to: str | None) -> tuple[int, str]:
    payload = {
        "personalizations": [{"to": [{"email": to_email}]}],
        "from": {"email": from_email, "name": from_name},
        "subject": subject,
        "content": [{"type": "text/plain", "value": body}],
        "tracking_settings": {
            "click_tracking": {"enable": False, "enable_text": False},
            "open_tracking": {"enable": False},
        },
    }
    if reply_to:
        payload["reply_to"] = {"email": reply_to}
    r = requests.post(
        SENDGRID_API_URL,
        headers={"Authorization": f"Bearer {api_key}",
                 "Content-Type": "application/json"},
        json=payload,
        timeout=20,
    )
    return r.status_code, (r.text or "")[:300]

# ─────────────────────────────────────────────────────────────────────────────
# Queue build (shared by --status / --dry-run / --send)
# ─────────────────────────────────────────────────────────────────────────────

def build_queue(today_product: str, batch_size: int, args, state: dict,
                logger: logging.Logger) -> dict:
    """Build today's queue. Side-effect: writes SendGrid suppression cache
    when called from --send. Raises SendGridSuppressionAbort or
    TemplateGuardFailed on fatal conditions."""

    # Primary-product template guard (runs in ALL modes — refuse to even preview
    # a bad template since dry-run could be screenshotted into review). Fill
    # pools are guarded individually below and skipped on failure, never fatal.
    templates = load_templates(today_product)
    violations = forbidden_token_check(today_product, templates)
    if violations:
        for v in violations:
            logger.error(f"FORBIDDEN_TOKEN: {v}")
        raise TemplateGuardFailed("Template token guard failed.")

    # Tracking + unsubscribe suppression
    touches, bounced, touch_count = load_tracking()
    unsub = load_unsubscribe_suppressions()
    if unsub:
        bounced = bounced | unsub
        logger.info(f"unsubscribe_csv: merged {len(unsub)} email(s) into bounced")

    sg_used = "none"
    sg_age = None
    sg_overlap: list[str] = []
    sg_counts: dict[str, int] = {}
    sg_emails: set[str] = set()

    if args.send:
        api_key = os.environ.get("SENDGRID_API_KEY")
        if not api_key:
            raise RuntimeError("SENDGRID_API_KEY not set in environment")
        sg_emails, sg_counts, n_failed = fetch_sendgrid_suppressions(api_key, logger)
        if n_failed == len(SENDGRID_SUPPRESSION_ENDPOINTS):
            cached_emails, cached_age, cached_counts = load_sendgrid_cache()
            if not cached_emails or cached_age is None or cached_age > SENDGRID_CACHE_MAX_AGE_HR:
                age_descr = "absent" if cached_age is None else f"stale ({cached_age:.1f}h)"
                raise SendGridSuppressionAbort(
                    f"All SendGrid suppression endpoints failed; cache {age_descr}."
                )
            sg_emails = cached_emails
            sg_counts = cached_counts
            sg_used = "cache"
            sg_age = cached_age
            logger.warning(f"sendgrid_suppressions: using cache ({sg_age:.1f}h old)")
        else:
            save_sendgrid_cache(sg_emails, sg_counts)
            sg_used = "fresh"
        if sg_emails:
            recent = _recent_send_emails(within_days=7)
            sg_overlap = sorted(sg_emails & recent)
            for em in sg_overlap:
                logger.warning(
                    f"sendgrid_suppression_overlap: {em} sent in last 7d "
                    f"(early reputation signal)"
                )
        logger.info(
            "sendgrid_suppressions: "
            f"bounces={sg_counts.get('bounces',0)} "
            f"blocks={sg_counts.get('blocks',0)} "
            f"spam={sg_counts.get('spam_reports',0)} "
            f"invalid={sg_counts.get('invalid_emails',0)} "
            f"unique={len(sg_emails)} cache={sg_used}"
            + (f" age={sg_age:.1f}h" if sg_age is not None else "")
        )
        bounced = bounced | sg_emails
    else:
        # --status / --dry-run: use cache only, no API call
        sg_emails, sg_age, sg_counts = load_sendgrid_cache()
        if sg_emails:
            bounced = bounced | sg_emails
            sg_used = "cache"
            logger.info(
                f"sendgrid_suppressions: cache used "
                f"({sg_age:.1f}h old, {len(sg_emails)} email(s))"
            )

    # ─────────────────────────────────────────────────────────────────────────
    # Multi-pool fill-to-cap. Historically the drip queued a SINGLE rotation
    # product; when that pool ran thin (coffee exhaustion, 7/9–7/11 shipped 2/1/1
    # against a 150 cap) it lost the day's capacity. Now we start at today's
    # rotation product and, if it can't fill `batch_size`, fall through to the
    # remaining products in rotation order — deduping by email AND domain across
    # pools — until the cap is met or every pool is drained. Each queued contact
    # carries its own product (`_product`) and rendered template (`_template`).
    # ─────────────────────────────────────────────────────────────────────────
    templates_by_product: dict[str, list[tuple[str, str, str]]] = {today_product: templates}

    def pool_eligible(product: str):
        """Load + country-filter + eligibility-filter + domain-throttle ONE
        product's pool. Returns (info_dict, templates) or (None, None) if the
        pool's templates fail the token guard — fill pools are skipped on guard
        failure (only the primary product, guarded above, aborts the run)."""
        tmpls = templates_by_product.get(product)
        if tmpls is None:
            tmpls = load_templates(product)
            v = forbidden_token_check(product, tmpls)
            if v:
                for vv in v:
                    logger.error(f"FORBIDDEN_TOKEN (fill pool {product}, skipped): {vv}")
                return None, None
            templates_by_product[product] = tmpls
        cands = load_xlsx_contacts(product, logger) \
              + load_named_contacts(product, logger) \
              + load_tracking_followup_contacts(product, logger)
        loaded = len(cands)
        cands, cexcl = apply_country_filter(cands, logger)
        elig = filter_eligible(cands, touches, bounced, touch_count, datetime.now())
        thr, dropped = domain_throttle(elig)
        return {
            "eligible": thr, "loaded": loaded, "country_excluded": cexcl,
            "eligible_total": len(elig), "domain_dropped": dropped,
        }, tmpls

    # Rotation order starting at today's product (e.g. coffee → sugar → pepper).
    order: list[str] = []
    _p = today_product
    for _ in range(len(PRODUCTS)):
        order.append(_p)
        _p = next_product(_p)

    queue: list[dict] = []
    seen_emails: set[str] = set()
    seen_domains: set[str] = set()
    pools_used: list[tuple[str, int]] = []
    candidates_loaded = country_excluded = eligible_total = domain_dropped = 0

    for product in order:
        if len(queue) >= batch_size:
            break
        info, tmpls = pool_eligible(product)
        if info is None:            # fill-pool template guard failed → skip pool
            continue
        candidates_loaded += info["loaded"]
        country_excluded  += info["country_excluded"]
        eligible_total    += info["eligible_total"]
        domain_dropped    += info["domain_dropped"]
        added = 0
        for c in info["eligible"]:
            if len(queue) >= batch_size:
                break
            em = c["email"].strip().lower()
            dom = em.split("@", 1)[1] if "@" in em else em
            if em in seen_emails or dom in seen_domains:   # cross-pool dedup
                continue
            seen_emails.add(em)
            seen_domains.add(dom)
            c["_product"] = product
            c["_template"] = tmpls[added % len(tmpls)]     # balance a/b/c per pool
            queue.append(c)
            added += 1
        if added:
            pools_used.append((product, added))
            logger.info(f"pool_fill: {product} +{added} (total {len(queue)}/{batch_size})")

    logger.info(
        f"queue: {len(queue)} (cap={batch_size}, "
        f"ramp_day={state.get('ramp_day', 0)}, pools={pools_used or 'none'})"
    )

    # Pool-exhaustion alert. The fill loop only stops short of `batch_size`
    # (which is always ≥10 under the ramp) when every pool combined is drained,
    # so a sub-threshold queue means the lead supply is effectively empty. Log
    # in every mode; email the admin only on --send (does NOT pause the drip).
    if len(queue) < POOL_EXHAUSTION_ALERT_THRESHOLD:
        logger.error(
            f"POOL_EXHAUSTION: only {len(queue)} eligible across all pools "
            f"{order} (threshold {POOL_EXHAUSTION_ALERT_THRESHOLD}) — refill leads"
        )
        if args.send:
            notify_pool_exhaustion(len(queue), pools_used, order, logger)

    # Tier breakdown over the final (post-fill) queue.
    tier_breakdown: dict[int, int] = {0: 0, 1: 0, 2: 0, 3: 0}
    for c in queue:
        t = int(c.get("tier", 3))
        tier_breakdown[t] = tier_breakdown.get(t, 0) + 1

    return {
        "queue": queue,
        "templates": templates,
        "pools_used": pools_used,
        "candidates_loaded": candidates_loaded,
        "country_excluded": country_excluded,
        "eligible_total": eligible_total,
        "tier_breakdown": tier_breakdown,
        "domain_dropped": domain_dropped,
        "queue_size": len(queue),
        "sendgrid_used": sg_used,
        "sendgrid_cache_age": sg_age,
        "sendgrid_overlap_with_recent_sends": sg_overlap,
        "sendgrid_counts": sg_counts,
        "sendgrid_unique": len(sg_emails),
    }

# ─────────────────────────────────────────────────────────────────────────────
# Modes
# ─────────────────────────────────────────────────────────────────────────────

def cmd_status(state: dict, today_product: str, batch_size: int,
               brake_rate: float, queue_info: dict | None) -> None:
    brake = state.get("safety_brake", {}) or {}
    scanner = get_scanner_health()
    print("=== CLADUTA DRIP — STATUS ===")
    if not scanner["healthy"]:
        print(f"  *** ALERT: Unsubscribe scanner {scanner['state']} — drip will refuse --send ***")
        for w in scanner["warnings"]:
            print(f"  ***   {w}")
        print()
    print(f"  Today's rotation product:  {today_product}")
    print(f"  Last product sent:         {state.get('last_product_sent') or '(none)'}")
    print(f"  Last send date:            {state.get('last_send_date') or '(never)'}")
    print(f"  Ramp day:                  {state.get('ramp_day', 0)}  (of {len(RAMP_SCHEDULE)} scheduled)")
    print(f"  Ramp schedule:             {RAMP_SCHEDULE}")
    print(f"  Today's batch ceiling:     {batch_size}  (hard cap {HARD_DAILY_CAP})")
    print(f"  Mid-run brake interval:    every {MID_RUN_BRAKE_INTERVAL} sends")
    print(f"  Max touches per email:     {MAX_TOUCHES_PER_EMAIL}")
    print(f"  7-day failure rate:        {brake_rate:.2f}%  (trip threshold {BOUNCE_RATE_TRIP_PCT}%)")
    print(f"  Safety brake active:       {brake.get('active')}")
    if brake.get("active"):
        print(f"    Reason:    {brake.get('reason')}")
        print(f"    Triggered: {brake.get('triggered_at')}")

    if queue_info is not None:
        tb = queue_info["tier_breakdown"]
        print()
        print("--- Today's queue preview ---")
        print(f"  Candidates loaded:         {queue_info['candidates_loaded']}")
        print(f"  Country-excluded:          {queue_info['country_excluded']}")
        print(f"  Eligible (post-filters):   {queue_info['eligible_total']}")
        print(f"  Tier breakdown:            T0={tb.get(0,0)}  T1={tb.get(1,0)}  T2={tb.get(2,0)}  T3={tb.get(3,0)}")
        print(f"  Domain-throttled removed:  {queue_info['domain_dropped']}")
        print(f"  Final queue size:          {queue_info['queue_size']}  (cap={batch_size})")

    print()
    print("--- SendGrid suppression cache ---")
    if queue_info is None or queue_info["sendgrid_used"] == "none":
        print("  (no cache yet; will populate on first --send)")
    else:
        sgc = queue_info["sendgrid_counts"]
        print(f"  Total unique emails:       {queue_info['sendgrid_unique']}")
        print(f"  Per-endpoint:              bounces={sgc.get('bounces',0)}  "
              f"blocks={sgc.get('blocks',0)}  spam={sgc.get('spam_reports',0)}  "
              f"invalid={sgc.get('invalid_emails',0)}")
        if queue_info["sendgrid_cache_age"] is not None:
            print(f"  Cache age:                 {queue_info['sendgrid_cache_age']:.1f}h")

    unsub = load_unsubscribe_suppressions()
    print()
    print("--- Unsubscribe suppression list ---")
    print(f"  Emails on list:            {len(unsub)}  ({UNSUBSCRIBE_CSV})")

    print()
    print("--- Unsubscribe scanner health ---")
    print(f"  Status:                    {scanner['state']}")
    if scanner.get("reason"):
        print(f"  Reason:                    {scanner['reason']}")
    if scanner.get("last_ok_age_hours") is not None:
        print(f"  Last OK age:               {scanner['last_ok_age_hours']:.1f}h"
              f"  (threshold {SCANNER_LAST_OK_MAX_AGE_HOURS}h)")
    print(f"  Will block --send:         {'YES' if not scanner['healthy'] else 'no'}")

    print()
    print(f"  Tracking CSV:              {TRACKING_CSV}")
    print(f"  Templates dir:             {TEMPLATES_DIR}")

def trip_safety_brake(state: dict, reason: str, now: datetime,
                      logger: logging.Logger) -> None:
    state["safety_brake"] = {
        "active": True,
        "reason": reason,
        "triggered_at": now.isoformat(),
    }
    state["ramp_day"] = 0
    save_state(state)

    api_key    = os.environ.get("SENDGRID_API_KEY")
    admin      = os.environ.get("DRIP_ADMIN_EMAIL", "jasondudney@gmail.com")
    from_email = os.environ.get("SENDGRID_FROM_EMAIL", "info@cladutacorp.com")
    from_name  = os.environ.get("SENDGRID_FROM_NAME",  "Jason Dudney")
    if not api_key:
        logger.warning("SENDGRID_API_KEY missing — cannot send safety-brake notification")
        return
    try:
        code, _text = send_via_sendgrid(
            api_key, from_email, from_name, admin,
            "[CLADUTA DRIP] Safety brake triggered — drip paused",
            (
                f"Drip paused at {now.isoformat()}.\n\n"
                f"Reason: {reason}\n\n"
                f"Ramp day has been reset to 0. The script will refuse to send\n"
                f"until you investigate and either:\n"
                f"  • Clear the brake by editing {STATE_FILE} (set "
                f"safety_brake.active = false)\n"
                f"  • Run with --batch N (still bounded by HARD_DAILY_CAP={HARD_DAILY_CAP})\n\n"
                f"Tracking CSV: {TRACKING_CSV}\n"
                f"Logs dir:     {LOGS_DIR}\n"
            ),
            from_email,
        )
        logger.info(f"safety-brake notification -> {admin}: HTTP {code}")
    except Exception as e:
        logger.warning(f"safety-brake notification send failed: {e}")

def notify_pool_exhaustion(queue_len: int, pools_used: list, order: list,
                           logger: logging.Logger) -> None:
    """Email the admin when every lead pool is effectively drained. Unlike the
    safety brake this does NOT pause the drip — it ships what little it found and
    asks for a lead refill."""
    api_key    = os.environ.get("SENDGRID_API_KEY")
    admin      = os.environ.get("DRIP_ADMIN_EMAIL", "jasondudney@gmail.com")
    from_email = os.environ.get("SENDGRID_FROM_EMAIL", "info@cladutacorp.com")
    from_name  = os.environ.get("SENDGRID_FROM_NAME",  "Jason Dudney")
    if not api_key:
        logger.warning("SENDGRID_API_KEY missing — cannot send pool-exhaustion notification")
        return
    try:
        code, _text = send_via_sendgrid(
            api_key, from_email, from_name, admin,
            "[CLADUTA DRIP] Lead pools exhausted — refill needed",
            (
                f"Today's drip could only queue {queue_len} eligible contact(s) "
                f"across all pools ({', '.join(order)}).\n\n"
                f"Per-pool contribution: {pools_used or 'none'}\n\n"
                f"The drip still sent what it found — it did NOT pause — but the\n"
                f"master lead list is effectively tapped. Refill leads to restore\n"
                f"capacity.\n\n"
                f"Tracking CSV: {TRACKING_CSV}\n"
                f"Logs dir:     {LOGS_DIR}\n"
            ),
            from_email,
        )
        logger.info(f"pool-exhaustion notification -> {admin}: HTTP {code}")
    except Exception as e:
        logger.warning(f"pool-exhaustion notification send failed: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> int:
    p = argparse.ArgumentParser(description="Claduta daily drip — rotating product cold email")
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--status",  action="store_true", help="show queue + state, no I/O")
    g.add_argument("--dry-run", action="store_true", help="show what WOULD send")
    g.add_argument("--send",    action="store_true", help="execute the send")
    p.add_argument("--batch", type=int, default=None,
                   help=f"override ramp; capped at {HARD_DAILY_CAP}")
    p.add_argument("--product", choices=PRODUCTS, default=None,
                   help="force a product (default: rotate from state)")
    args = p.parse_args()

    now = datetime.now()
    run_id = now.strftime("%Y%m%d_%H%M")
    logger = setup_logger(run_id)
    logger.info(f"=== claduta_drip start (run_id={run_id}) ===")

    state = load_state()

    # BUG 4: ramp gap reset
    last_send_date_str = state.get("last_send_date")
    if last_send_date_str:
        try:
            last_send_d = datetime.strptime(last_send_date_str, "%Y-%m-%d").date()
            days_since = (now.date() - last_send_d).days
            if days_since > 3 and state.get("ramp_day", 0) > 0:
                logger.info(
                    f"Ramp gap: {days_since} days since last send "
                    f"({last_send_date_str}). Resetting ramp_day "
                    f"{state.get('ramp_day', 0)} -> 0."
                )
                state["ramp_day"] = 0
        except ValueError:
            logger.warning(f"Could not parse last_send_date={last_send_date_str!r}")

    today_product = args.product or next_product(state.get("last_product_sent"))

    ramp_size = ramp_batch_size(state.get("ramp_day", 0))
    if args.batch is not None:
        batch_size = max(1, min(args.batch, HARD_DAILY_CAP))
    else:
        batch_size = min(ramp_size, HARD_DAILY_CAP)

    brake_rate = compute_7d_failure_rate()

    # BLOCKER 2: surface scanner-health warnings in every mode. --send is
    # additionally gated below (refuses if unhealthy/stale); --status and
    # --dry-run continue but emit a loud warning at top of the run log.
    scanner = get_scanner_health()
    if not scanner["healthy"]:
        for w in scanner["warnings"]:
            logger.warning(f"SCANNER_HEALTH: {w}")

    # ─── --status ───────────────────────────────────────────────────────────
    if args.status:
        try:
            qi = build_queue(today_product, batch_size, args, state, logger)
        except (TemplateGuardFailed, SendGridSuppressionAbort, RuntimeError) as e:
            logger.error(str(e))
            qi = None
        cmd_status(state, today_product, batch_size, brake_rate, qi)
        return 0

    # Honor any already-tripped brake (any mode that would write)
    brake = state.get("safety_brake", {}) or {}
    if brake.get("active"):
        logger.error(
            f"SAFETY_BRAKE_ACTIVE — reason: {brake.get('reason')}; "
            f"triggered at {brake.get('triggered_at')}. Aborting."
        )
        return 3

    # Trip the brake on a fresh failure-rate spike
    if brake_rate > BOUNCE_RATE_TRIP_PCT:
        msg = f"7-day failure rate {brake_rate:.2f}% > {BOUNCE_RATE_TRIP_PCT}% threshold"
        logger.error(f"SAFETY_BRAKE_TRIGGERED: {msg}")
        trip_safety_brake(state, msg, now, logger)
        return 3

    # ─── --dry-run ──────────────────────────────────────────────────────────
    if args.dry_run:
        try:
            qi = build_queue(today_product, batch_size, args, state, logger)
        except TemplateGuardFailed:
            return 2
        except SendGridSuppressionAbort as e:
            logger.error(str(e))
            return 6
        except RuntimeError as e:
            logger.error(str(e))
            return 2
        queue = qi["queue"]
        touches, _, _ = load_tracking()
        logger.info(f"=== DRY RUN (no SendGrid calls, no CSV writes) pools={qi['pools_used']} ===")
        for i, c in enumerate(queue, 1):
            vid, subj, body = c["_template"]
            rsubj, _rbody = render(subj, body, c)
            label = "NEW" if c["email"].strip().lower() not in touches else "RE-ENGAGE"
            logger.info(
                f"[{i:>3}/{len(queue)}] {label:9} {c['_product']:6} T{c.get('tier',3)} {vid}  "
                f"{c['email']:<45}  {c['company']}"
            )
        if queue:
            example_v, example_subj, example_body = queue[0]["_template"]
            logger.info(f"--- example render ({queue[0]['_product']}/{example_v}) ---")
            ex_subj, ex_body = render(example_subj, example_body, queue[0])
            logger.info(f"Subject: {ex_subj}")
            logger.info("Body:\n" + ex_body)
        return 0

    # ─── --send ─────────────────────────────────────────────────────────────
    # BLOCKER 2: refuse to send when the unsubscribe scanner is unhealthy.
    if not scanner["healthy"]:
        logger.error(
            f"Refusing --send: unsubscribe scanner state={scanner['state']}. "
            f"Run process_unsubscribes.py to recover, or investigate "
            f"{SCANNER_UNHEALTHY_FILE} / {SCANNER_LAST_OK_FILE}."
        )
        return 7

    if not acquire_send_lock(logger):
        return 5

    try:
        qi = build_queue(today_product, batch_size, args, state, logger)
    except TemplateGuardFailed:
        return 2
    except SendGridSuppressionAbort as e:
        logger.error(str(e))
        return 6
    except RuntimeError as e:
        logger.error(str(e))
        return 4

    queue = qi["queue"]

    api_key    = os.environ.get("SENDGRID_API_KEY")
    from_email = os.environ.get("SENDGRID_FROM_EMAIL", "info@cladutacorp.com")
    from_name  = os.environ.get("SENDGRID_FROM_NAME",  "Jason Dudney")
    reply_to   = os.environ.get("SENDGRID_REPLY_TO",   from_email)

    # ITEM F: idempotency pre-build — anything sent in last 24h with the same
    # (email, product, date) tuple gets skipped before the SendGrid call.
    recent_keys = load_recent_send_keys(within_hours=24)
    today_iso = now.strftime("%Y-%m-%d")

    succeeded = rejected = api_errors = skipped = 0
    brake_tripped_mid_run = False
    logger.info(
        f"=== LIVE SEND  primary={today_product}  pools={qi['pools_used']}  "
        f"cap={batch_size} ==="
    )
    for i, c in enumerate(queue, 1):
        prod = c["_product"]
        em_lower = c["email"].strip().lower()
        key = (em_lower, prod, today_iso)
        if key in recent_keys:
            skipped += 1
            logger.info(
                f"[{i:>3}/{len(queue)}] IDEMPOTENT_SKIP: {c['email']} "
                f"already sent {prod} today"
            )
            continue

        send_ts = datetime.now()
        vid, subj, body = c["_template"]
        rsubj, rbody = render(subj, body, c)
        logger.info(
            f"[{i:>3}/{len(queue)}] -> {c['email']}  "
            f"({c['company']}, T{c.get('tier',3)}, {vid})"
        )
        try:
            code, text = send_via_sendgrid(
                api_key, from_email, from_name,
                c["email"], rsubj, rbody, reply_to,
            )
        except Exception as e:
            code, text = 0, str(e)
        # BLOCKER 4: distinguish real recipient rejections from transient API
        # failures so the bounce rate reflects reputation, not infra health.
        if 200 <= code < 300:
            status = "SENT"
            succeeded += 1
            logger.info(f"    HTTP {code}  SENT")
        elif code == 429:
            status = "API_THROTTLED"
            api_errors += 1
            logger.info(f"    HTTP {code}  API_THROTTLED  body: {text}")
        elif 400 <= code < 500:
            status = "REJECTED"
            rejected += 1
            logger.info(f"    HTTP {code}  REJECTED  body: {text}")
        elif code == 0 or code >= 500:
            status = "API_ERROR"
            api_errors += 1
            logger.info(f"    HTTP {code}  API_ERROR  body: {text}")
        else:
            status = "API_ERROR"
            api_errors += 1
            logger.info(f"    HTTP {code}  API_ERROR (unexpected code)  body: {text}")
        append_tracking_row({
            "email":   c["email"],
            "company": c["company"],
            "product": prod,
            "city":    c.get("city", ""),
            "sent_at": send_ts.strftime("%Y-%m-%dT%H:%M:%S"),
            "status":  status,
        })
        recent_keys.add(key)

        # Mid-run brake (every MID_RUN_BRAKE_INTERVAL = 10 sends)
        if i % MID_RUN_BRAKE_INTERVAL == 0 and i < len(queue):
            this_run_failed = rejected + api_errors
            this_run_rate = (this_run_failed / (succeeded + this_run_failed) * 100.0) if (succeeded + this_run_failed) else 0.0
            mid_run_rate = compute_7d_failure_rate()
            logger.info(
                f"    mid-run check at {i}/{len(queue)}: "
                f"this-run failure {this_run_rate:.1f}% | "
                f"7d failure {mid_run_rate:.2f}%"
            )
            if mid_run_rate > BOUNCE_RATE_TRIP_PCT:
                msg = (
                    f"Mid-run trip at send {i}/{len(queue)}: "
                    f"7-day failure rate {mid_run_rate:.2f}% > "
                    f"{BOUNCE_RATE_TRIP_PCT}% threshold"
                )
                logger.error(f"SAFETY_BRAKE_TRIGGERED (mid-run): {msg}")
                trip_safety_brake(state, msg, datetime.now(), logger)
                brake_tripped_mid_run = True
                logger.info(f"Aborting cleanly after {i} of {len(queue)} sends.")
                break

        if i < len(queue):
            logger.info(f"    sleeping {SEND_PACING_SECONDS}s...")
            time.sleep(SEND_PACING_SECONDS)

    total_failed = rejected + api_errors  # backward-compat "failed" total
    logger.info(
        f"=== RESULTS === Sent: {succeeded} | Rejected: {rejected} | "
        f"API errors: {api_errors} | Skipped (idempotent): {skipped}"
        + (" | mid-run brake tripped" if brake_tripped_mid_run else "")
    )

    append_summary_row({
        "date":            now.strftime("%Y-%m-%d"),
        "product":         today_product,
        "attempted":       len(queue),
        "succeeded":       succeeded,
        "rejected":        rejected,
        "api_errors":      api_errors,
        "failed":          total_failed,
        "bounce_rate_7d":  f"{compute_7d_failure_rate():.2f}",
    })

    # Advance state only if at least one send landed; don't bump ramp_day
    # when the brake tripped mid-run (trip_safety_brake reset it to 0).
    if succeeded > 0:
        state["last_product_sent"] = today_product
        state["last_send_date"]    = now.strftime("%Y-%m-%d")
        if not brake_tripped_mid_run:
            state["ramp_day"] = state.get("ramp_day", 0) + 1
    save_state(state)

    if brake_tripped_mid_run:
        return 3
    return 0 if (rejected == 0 and api_errors == 0) else 1

if __name__ == "__main__":
    sys.exit(main())
